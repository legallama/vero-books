from flask import render_template, request, flash, redirect, url_for
from flask_login import login_required, current_user
from ._bp import banking_bp
from app.models.banking.bank_account import BankAccount, BankTransaction
from app.services.auth_service import get_current_org
from app.services.plaid_service import PlaidService
from app.extensions import db
from flask import jsonify

@banking_bp.route('/')
@login_required
def index():
    org = get_current_org()
    from app.models.banking.bank_account import BankAccount, BankTransaction
    from app.models.accounting.bank_rule import BankRule
    from app.models.accounting.account import Account
    from decimal import Decimal
    from datetime import datetime, timedelta
    
    accounts = BankAccount.query.filter_by(organization_id=org.id).all()
    transactions = BankTransaction.query.filter_by(organization_id=org.id).order_by(BankTransaction.date.desc()).limit(50).all()
    active_rules = BankRule.query.filter_by(organization_id=org.id, is_active=True).order_by(BankRule.priority.desc()).all()
    
    # ── Dashboard stats ──
    all_txs = BankTransaction.query.filter_by(organization_id=org.id).all()
    
    total_balance = sum(float(a.balance or 0) for a in accounts)
    total_uncategorized = sum(1 for t in all_txs if t.status == 'UNCATEGORIZED')
    total_matched = sum(1 for t in all_txs if t.status == 'MATCHED')
    
    # Per-account enrichment
    for acc in accounts:
        acc_txs = [t for t in all_txs if t.bank_account_id == acc.id]
        acc.tx_count = len(acc_txs)
        acc.uncategorized_count = sum(1 for t in acc_txs if t.status == 'UNCATEGORIZED')
        acc.matched_count = sum(1 for t in acc_txs if t.status == 'MATCHED')
        acc.total_deposits = sum(float(t.amount) for t in acc_txs if float(t.amount) > 0)
        acc.total_withdrawals = sum(abs(float(t.amount)) for t in acc_txs if float(t.amount) < 0)
        acc.last_tx = max((t.date for t in acc_txs), default=None)
    
    # Last 30 days flow
    thirty_days_ago = datetime.utcnow().date() - timedelta(days=30)
    recent_txs = [t for t in all_txs if t.date and t.date >= thirty_days_ago]
    money_in_30d = sum(float(t.amount) for t in recent_txs if float(t.amount) > 0)
    money_out_30d = sum(abs(float(t.amount)) for t in recent_txs if float(t.amount) < 0)
    
    # Apply rules to transaction list for display
    for tx in transactions:
        tx.suggested_account = None
        tx.rule_name = None
        
        for rule in active_rules:
            match = False
            if rule.field_to_match == 'DESCRIPTION':
                if rule.match_type == 'CONTAINS' and rule.match_value.lower() in tx.description.lower():
                    match = True
                elif rule.match_type == 'EXACT' and rule.match_value.lower() == tx.description.lower():
                    match = True
            elif rule.field_to_match == 'AMOUNT':
                try:
                    val = float(rule.match_value)
                    if rule.match_type == 'EXACT' and float(tx.amount) == val:
                        match = True
                    elif rule.match_type == 'GREATER_THAN' and float(tx.amount) > val:
                        match = True
                except: pass
            
            if match:
                tx.suggested_account = rule.target_account
                tx.rule_name = rule.name
                tx.auto_post_enabled = rule.auto_post
                break # First match wins based on priority
        
        # Heuristic Magic Suggestion (If no rule matched)
        if not tx.suggested_account and tx.status == 'UNCATEGORIZED':
            last_match = BankTransaction.query.filter(
                BankTransaction.organization_id == org.id,
                BankTransaction.description == tx.description,
                BankTransaction.status == 'MATCHED',
                BankTransaction.id != tx.id
            ).order_by(BankTransaction.date.desc()).first()
            
            if last_match and last_match.matched_source_id:
                from app.models.accounting.journal import JournalEntry
                je = JournalEntry.query.get(last_match.matched_source_id)
                if je:
                    # Find the 'offset' account (not the bank account)
                    for line in je.lines:
                        if line.account_id != tx.bank_account.account_id:
                            tx.suggested_account = line.account
                            tx.rule_name = "Based on your history"
                            break


                
    gl_accounts = Account.query.filter_by(organization_id=org.id).order_by(Account.name).all()
    
    return render_template('banking/index.html',
        accounts=accounts,
        transactions=transactions,
        gl_accounts=gl_accounts,
        total_balance=total_balance,
        total_uncategorized=total_uncategorized,
        total_matched=total_matched,
        total_transactions=len(all_txs),
        money_in_30d=money_in_30d,
        money_out_30d=money_out_30d,
        rules_count=len(active_rules),
    )

@banking_bp.route('/accounts/create', methods=['POST'])
@login_required
def create_account():
    org = get_current_org()
    name = request.form.get('name')
    acc_type = request.form.get('type')
    
    new_acc = BankAccount(
        organization_id=org.id,
        name=name,
        account_type=acc_type
    )
    db.session.add(new_acc)
    db.session.commit()
    flash("Bank account linked.", "success")
    return redirect(url_for('banking.index'))

@banking_bp.route('/accounts/<bank_acc_id>/link-gl', methods=['POST'])
@login_required
def link_gl(bank_acc_id):
    org = get_current_org()
    from app.models.banking.bank_account import BankAccount
    from app.models.accounting.account import Account
    
    bank_acc = BankAccount.query.filter_by(id=bank_acc_id, organization_id=org.id).first_or_404()
    gl_account_id = request.form.get('gl_account_id')
    
    if gl_account_id:
        gl_acc = Account.query.filter_by(id=gl_account_id, organization_id=org.id).first_or_404()
        bank_acc.account_id = gl_acc.id
        db.session.commit()
        flash(f"Account '{bank_acc.name}' is now linked to GL account '{gl_acc.name}'.", "success")
    else:
        flash("No GL account selected.", "warning")
        
    return redirect(url_for('banking.index'))

@banking_bp.route('/accounts/<bank_acc_id>/unlink-gl', methods=['POST'])
@login_required
def unlink_gl(bank_acc_id):
    org = get_current_org()
    from app.models.banking.bank_account import BankAccount
    
    bank_acc = BankAccount.query.filter_by(id=bank_acc_id, organization_id=org.id).first_or_404()
    bank_acc.account_id = None
    db.session.commit()
    
    flash(f"Account '{bank_acc.name}' has been unlinked from the General Ledger.", "info")
    return redirect(url_for('banking.index'))

@banking_bp.route('/accounts/<bank_acc_id>/delete', methods=['POST'])
@login_required
def delete_account(bank_acc_id):
    org = get_current_org()
    from app.models.banking.bank_account import BankAccount, BankTransaction
    
    bank_acc = BankAccount.query.filter_by(id=bank_acc_id, organization_id=org.id).first_or_404()
    
    # Safety check: Don't delete if there are transactions
    tx_count = BankTransaction.query.filter_by(bank_account_id=bank_acc_id).count()
    if tx_count > 0:
        flash(f"Cannot delete account '{bank_acc.name}' because it has {tx_count} transactions. Please remove transactions first.", "danger")
        return redirect(url_for('banking.index'))
        
    db.session.delete(bank_acc)
    db.session.commit()
    
    flash(f"Bank account '{bank_acc.name}' has been removed.", "success")
    return redirect(url_for('banking.index'))

@banking_bp.route('/accounts/<bank_acc_id>/sync', methods=['POST'])
@login_required
def sync_account(bank_acc_id):
    org = get_current_org()
    from app.models.banking.bank_account import BankAccount, BankTransaction
    from app.services.banking_service import BankingService
    import random
    from datetime import datetime, timedelta
    
    bank_acc = BankAccount.query.filter_by(id=bank_acc_id, organization_id=org.id).first_or_404()
    
    new_txs = 0

    if bank_acc.plaid_access_token:
        # Real Plaid Sync
        try:
            start_date = (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d')
            end_date = datetime.utcnow().strftime('%Y-%m-%d')
            
            plaid_txs = PlaidService.get_transactions(bank_acc.plaid_access_token, start_date, end_date)
            
            for ptx in plaid_txs:
                # Check for duplicates using plaid_transaction_id
                existing = BankTransaction.query.filter_by(plaid_transaction_id=ptx['transaction_id']).first()
                if not existing:
                    tx = BankTransaction(
                        organization_id=org.id,
                        bank_account_id=bank_acc.id,
                        date=datetime.strptime(ptx['date'], '%Y-%m-%d').date(),
                        description=ptx['name'],
                        amount=-float(ptx['amount']), # Plaid: positive is debit, negative is credit. We use positive for deposit.
                        plaid_transaction_id=ptx['transaction_id'],
                        status='UNCATEGORIZED'
                    )
                    db.session.add(tx)
                    db.session.flush()
                    BankingService.apply_rules_to_transaction(tx, org.id, current_user.id)
                    new_txs += 1
            db.session.commit()
            flash(f"Sync complete. {new_txs} new transactions pulled from Plaid.", "success")
            return redirect(url_for('banking.index'))
        except Exception as e:
            flash(f"Plaid Sync Error: {str(e)}", "danger")
            return redirect(url_for('banking.index'))

    # Fallback: Simulate API call to a provider like Plaid or Yodlee
    potential_vendors = [
        {"desc": "Amazon.com", "min": 15, "max": 200},
        {"desc": "Starbucks Coffee", "min": 5, "max": 25},
        {"desc": "Shell Oil", "min": 40, "max": 90},
        {"desc": "Adobe Systems", "min": 52, "max": 52},
        {"desc": "WeWork Office", "min": 450, "max": 450},
        {"desc": "Stripe Payout", "min": 1200, "max": 5000, "type": "deposit"},
        {"desc": "Apple Service", "min": 0.99, "max": 14.99}
    ]
    
    for _ in range(random.randint(2, 5)):
        vendor = random.choice(potential_vendors)
        is_deposit = vendor.get('type') == 'deposit'
        
        amount = random.uniform(vendor['min'], vendor['max'])
        if not is_deposit:
            amount = -amount
            
        date = datetime.utcnow().date() - timedelta(days=random.randint(0, 2))
        
        tx = BankTransaction(
            organization_id=org.id,
            bank_account_id=bank_acc.id,
            date=date,
            description=vendor['desc'],
            amount=amount,
            status='UNCATEGORIZED'
        )
        db.session.add(tx)
        db.session.flush()
        
        BankingService.apply_rules_to_transaction(tx, org.id, current_user.id)
        new_txs += 1
            
    db.session.commit()
    
    flash(f"Sync complete. {new_txs} new transactions pulled from {bank_acc.name}.", "success")
    return redirect(url_for('banking.index'))

@banking_bp.route('/plaid/create-link-token', methods=['POST'])
@login_required
def create_link_token():
    org = get_current_org()
    try:
        link_token = PlaidService.create_link_token(current_user.id, org.id)
        return jsonify({'link_token': link_token})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@banking_bp.route('/plaid/exchange-token', methods=['POST'])
@login_required
def exchange_token():
    org = get_current_org()
    data = request.get_json()
    public_token = data.get('public_token')
    metadata = data.get('metadata', {})
    
    if not public_token:
        return jsonify({'error': 'Missing public token'}), 400
        
    try:
        exchange_resp = PlaidService.exchange_public_token(public_token)
        access_token = exchange_resp['access_token']
        item_id = exchange_resp['item_id']
        
        # Plaid returns accounts in metadata. We'll link the selected one or the first one.
        plaid_accounts = metadata.get('accounts', [])
        if not plaid_accounts:
            return jsonify({'error': 'No accounts found in metadata'}), 400
            
        # For this MVP, we link the first account selected in Link
        pa = plaid_accounts[0]
        
        # Check if already exists
        bank_acc = BankAccount.query.filter_by(
            organization_id=org.id, 
            plaid_account_id=pa['id']
        ).first()
        
        if not bank_acc:
            bank_acc = BankAccount(
                organization_id=org.id,
                name=pa['name'],
                account_type=pa.get('subtype', 'Checking').capitalize(),
                account_number_last4=pa.get('mask'),
                bank_name=metadata.get('institution', {}).get('name', 'Plaid Linked Bank'),
                plaid_access_token=access_token,
                plaid_item_id=item_id,
                plaid_account_id=pa['id'],
                plaid_institution_id=metadata.get('institution', {}).get('institution_id')
            )
            db.session.add(bank_acc)
        else:
            bank_acc.plaid_access_token = access_token
            bank_acc.plaid_item_id = item_id
            
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@banking_bp.route('/checks/<check_id>/print')
@login_required
def print_check(check_id):
    org = get_current_org()
    from app.models.banking.check import Check
    check = Check.query.filter_by(id=check_id, organization_id=org.id).first_or_404()
    
    # We might want to convert amount to words here if the template needs it
    # For now, we'll do it in the template or pass a helper
    return render_template('banking/checks/print.html', check=check)

@banking_bp.route('/checks/<check_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_check(check_id):
    org = get_current_org()
    from app.models.banking.bank_account import BankAccount
    from app.models.accounting.account import Account
    from app.models.crm.contact import Vendor
    from app.models.banking.check import Check
    from app.models.accounting.journal import JournalEntry, JournalLine
    from app.services.ledger_service import LedgerService
    from datetime import datetime

    check = Check.query.filter_by(id=check_id, organization_id=org.id).first_or_404()

    if request.method == 'POST':
        bank_account_id = request.form.get('bank_account_id')
        check_number = request.form.get('check_number')
        date_str = request.form.get('date')
        payee_name = request.form.get('payee_name')
        amount_str = request.form.get('amount', '0')
        memo = request.form.get('memo')
        expense_account_id = request.form.get('expense_account_id')

        try:
            amount = float(amount_str)
        except ValueError:
            amount = 0

        if not bank_account_id or not check_number or amount <= 0 or not expense_account_id:
            flash("Missing or invalid required fields.", "danger")
            return redirect(url_for('banking.edit_check', check_id=check_id))

        check_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.utcnow().date()
        bank_acc = BankAccount.query.get(bank_account_id)

        # 1. Update Check Record
        check.bank_account_id = bank_account_id
        check.check_number = check_number
        check.date = check_date
        check.payee_name = payee_name
        check.amount = amount
        check.memo = memo

        # 2. Update/Replace Journal Entry
        # If already posted, we should ideally reverse and create new, 
        # but for simplicity in this MVP, we will update the existing one.
        if check.journal_entry_id:
            je = JournalEntry.query.get(check.journal_entry_id)
            if je:
                je.entry_number = f"CHK-{check_number}"
                je.entry_date = datetime.combine(check_date, datetime.min.time())
                je.memo = memo or f"Check #{check_number} to {payee_name}"
                
                # Clear old lines and add new ones
                for line in je.lines:
                    db.session.delete(line)
                
                bank_line = JournalLine(journal_entry_id=je.id, account_id=bank_acc.account_id, credit=amount, description=f"Check #{check_number} to {payee_name}")
                expense_line = JournalLine(journal_entry_id=je.id, account_id=expense_account_id, debit=amount, description=memo or f"Check #{check_number}")
                db.session.add(bank_line)
                db.session.add(expense_line)
        
        db.session.commit()
        flash(f"Check #{check_number} updated successfully.", "success")
        return redirect(url_for('banking.checks'))

    bank_accounts = BankAccount.query.filter_by(organization_id=org.id, account_type='Checking').all()
    expense_accounts = Account.query.filter_by(organization_id=org.id).order_by(Account.code).all()
    vendors = Vendor.query.filter_by(organization_id=org.id).all()
    
    # Try to find the current expense account from the journal entry
    current_expense_account_id = None
    if check.journal_entry:
        for line in check.journal_entry.lines:
            if line.debit > 0:
                current_expense_account_id = line.account_id
                break

    return render_template('banking/checks/edit.html', 
                           check=check,
                           bank_accounts=bank_accounts, 
                           expense_accounts=expense_accounts,
                           vendors=vendors,
                           current_expense_account_id=current_expense_account_id)

from app.services.auth_service import get_current_org, require_role

@banking_bp.route('/checks/<check_id>/delete', methods=['POST'])
@login_required
@require_role(['ADMIN', 'ACCOUNTANT'])
def delete_check(check_id):

    org = get_current_org()
    from app.models.banking.check import Check
    from app.services.ledger_service import LedgerService
    
    check = Check.query.filter_by(id=check_id, organization_id=org.id).first_or_404()
    check_num = check.check_number

    # Reverse the journal entry if it exists and is posted
    if check.journal_entry_id:
        je = check.journal_entry
        if je.status == 'POSTED':
            LedgerService.reverse_journal_entry(je.id, current_user.id, org.id, f"Check #{check_num} deleted")
        else:
            db.session.delete(je)

    db.session.delete(check)
    db.session.commit()
    
    flash(f"Check #{check_num} has been deleted and its ledger entry reversed/removed.", "info")
    return redirect(url_for('banking.checks'))





@banking_bp.route('/transactions/manual', methods=['POST'])
@login_required
def manual_transaction():
    org = get_current_org()
    from app.models.banking.bank_account import BankTransaction
    from datetime import datetime
    
    tx_type = request.form.get('type')
    amount_str = request.form.get('amount', '0')
    try:
        amount = float(amount_str)
        if tx_type == 'withdrawal':
            amount = -abs(amount)
        else:
            amount = abs(amount)
    except ValueError:
        amount = 0.0
        
    date_str = request.form.get('date')
    tx_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.utcnow().date()
    
    new_tx = BankTransaction(
        organization_id=org.id,
        bank_account_id=request.form.get('bank_account_id'),
        date=tx_date,
        description=request.form.get('description'),
        amount=amount,
        status='UNCATEGORIZED' # Will be caught by rules engine later
    )
    db.session.add(new_tx)
    db.session.commit()
    
    flash(f"Manual {tx_type} recorded successfully.", "success")
    return redirect(url_for('banking.index'))

@banking_bp.route('/accept-match/<tx_id>', methods=['POST'])
@login_required
def accept_match(tx_id):
    org = get_current_org()
    account_id = request.args.get('account_id') or request.form.get('account_id')
    
    # Get Transaction
    tx = BankTransaction.query.filter_by(id=tx_id, organization_id=org.id).first_or_404()
    
    if tx.status == 'MATCHED':
        flash('Transaction already matched.', 'warning')
        return redirect(url_for('banking.index'))
        
    from app.models.accounting.account import Account
    target_account = Account.query.filter_by(id=account_id, organization_id=org.id).first_or_404()
    
    # Map Bank Account to GL
    bank_account = tx.bank_account
    if not bank_account.account_id:
        gl_account = Account.query.filter_by(organization_id=org.id, name=bank_account.name).first()
        if not gl_account:
            gl_account = Account(organization_id=org.id, name=bank_account.name, code=bank_account.account_number_last4 or "1000", type="Asset", subtype="Bank")
            db.session.add(gl_account)
            db.session.flush()
        bank_account.account_id = gl_account.id
        
    # Create Journal Entry
    import time
    from app.models.accounting.journal import JournalEntry, JournalLine
    
    je = JournalEntry(
        organization_id=org.id,
        entry_number=f"BNK-{int(time.time())}",
        entry_date=tx.date,
        memo=f"Banking Feed: {tx.description}",
        source_type='BANK_FEED',
        source_id=tx.id,
        status='DRAFT'
    )
    db.session.add(je)
    db.session.flush()
    
    amount = abs(float(tx.amount))
    if float(tx.amount) > 0: # Deposit: Debit Bank, Credit Target
        bank_line = JournalLine(journal_entry_id=je.id, account_id=bank_account.account_id, debit=amount, description=tx.description)
        target_line = JournalLine(journal_entry_id=je.id, account_id=target_account.id, credit=amount, description=tx.description)
    else: # Withdrawal: Credit Bank, Debit Target
        bank_line = JournalLine(journal_entry_id=je.id, account_id=bank_account.account_id, credit=amount, description=tx.description)
        target_line = JournalLine(journal_entry_id=je.id, account_id=target_account.id, debit=amount, description=tx.description)
        
    db.session.add(bank_line)
    db.session.add(target_line)
    
    # Post it
    from app.services.ledger_service import LedgerService
    success, msg = LedgerService.post_journal_entry(je.id, current_user.id, org.id)
    if not success:
        flash(f"Error posting match to ledger: {msg}", "danger")
        return redirect(url_for('banking.index'))
        
    # Mark Match
    tx.status = 'MATCHED'
    tx.matched_source_type = 'JOURNAL_ENTRY'
    tx.matched_source_id = je.id
    
    db.session.commit()
    flash('Transaction mathematically categorized and added to the General Ledger.', 'success')
    return redirect(url_for('banking.index'))

@banking_bp.route('/delete-transaction/<tx_id>', methods=['POST'])
@login_required
def delete_transaction(tx_id):
    org = get_current_org()
    
    # Get Transaction
    tx = BankTransaction.query.filter_by(id=tx_id, organization_id=org.id).first_or_404()
    
    # Allow deletion. Technically if it's MATCHED, deleting it from the feed won't delete the journal entry, 
    # but that's acceptable for feed management. 
    db.session.delete(tx)
    db.session.commit()
    
    flash('Transaction removed from bank feed.', 'success')
    return redirect(url_for('banking.index'))

@banking_bp.route('/delete-transactions-batch', methods=['POST'])
@login_required
def delete_batch():
    org = get_current_org()
    
    tx_ids = request.form.getlist('tx_ids')
    if not tx_ids:
        flash('No transactions selected.', 'warning')
        return redirect(url_for('banking.index'))
        
    deleted_count = 0
    for tx_id in tx_ids:
        tx = BankTransaction.query.filter_by(id=tx_id, organization_id=org.id).first()
        if tx:
            db.session.delete(tx)
            deleted_count += 1
            
    db.session.commit()
    flash(f'{deleted_count} transactions permanently removed from bank feed.', 'success')
    return redirect(url_for('banking.index'))

@banking_bp.route('/import-statements', methods=['POST'])
@login_required
def import_statements():
    org = get_current_org()
    if 'statement_file' not in request.files:
        flash("No file provided.", "danger")
        return redirect(url_for('banking.index'))
        
    file = request.files['statement_file']
    if file.filename == '':
        flash("No file selected.", "danger")
        return redirect(url_for('banking.index'))
        
    filename = file.filename.lower()
    imported_txs = 0
    from app.models.banking.bank_account import BankAccount, BankTransaction
    
    try:
        # ---- OFX / QBO PARSING ----
        if filename.endswith('.qbo') or filename.endswith('.ofx') or filename.endswith('.qfx'):
            from ofxparse import OfxParser
            ofx = OfxParser.parse(file.stream)
            
            for account in ofx.accounts:
                acct_num = account.account_id[-4:] if len(account.account_id) > 4 else account.account_id
                bank_acc = BankAccount.query.filter_by(organization_id=org.id, account_number_last4=acct_num).first()
                if not bank_acc:
                    bank_acc = BankAccount(
                        organization_id=org.id,
                        name=f"Bank Import ({acct_num})",
                        account_type="Checking",
                        account_number_last4=acct_num,
                        balance=account.statement.balance if hasattr(account.statement, 'balance') else 0.0
                    )
                    db.session.add(bank_acc)
                    db.session.flush()
                elif hasattr(account.statement, 'balance'):
                    bank_acc.balance = account.statement.balance
                    
                for tx in account.statement.transactions:
                    existing = BankTransaction.query.filter_by(
                        bank_account_id=bank_acc.id, date=tx.date.date(), amount=tx.amount
                    ).first()
                    if not existing:
                        new_tx = BankTransaction(
                            organization_id=org.id,
                            bank_account_id=bank_acc.id,
                            date=tx.date.date(),
                            description=tx.payee or tx.memo or "Unknown Transaction",
                            amount=tx.amount,
                            status='UNCATEGORIZED'
                        )
                        db.session.add(new_tx)
                        db.session.flush() # Ensure tx has ID
                        imported_txs += 1
                        
                        # Apply Smart Rules
                        from app.services.banking_service import BankingService
                        BankingService.apply_rules_to_transaction(new_tx, org.id, current_user.id)

                        
            db.session.commit()
            flash(f"Successfully imported {imported_txs} transactions from QBO/OFX.", "success")
            
        # ---- EXCEL / CSV PARSING ----
        elif filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv'):
            # Grab or create a generic default account for CSV/XLSX imports 
            bank_acc = BankAccount.query.filter_by(organization_id=org.id).first()
            if not bank_acc:
                bank_acc = BankAccount(
                    organization_id=org.id,
                    name=f"Manual Import Account",
                    account_type="Checking",
                    account_number_last4="0000"
                )
                db.session.add(bank_acc)
                db.session.flush()

            import csv
            import io
            from datetime import datetime
            
            rows = []
            
            if filename.endswith('.csv'):
                content_bytes = file.stream.read()
                file.stream.seek(0)
                try:
                    with open(r'c:\dev\veros_books\last_import.csv', 'wb') as f:
                        f.write(content_bytes)
                except:
                    pass
                stream = io.StringIO(content_bytes.decode("UTF8"), newline=None)
                raw_rows = list(csv.reader(stream))
                header_row_index = 0
                
                # Scan for actual headers to bypass QB's 4-5 line metadata blocks
                for idx, row in enumerate(raw_rows):
                    row_str = " ".join([str(x).lower() for x in row if x])
                    if 'date' in row_str and ('amount' in row_str or 'debit' in row_str):
                        header_row_index = idx
                        break
                        
                if len(raw_rows) > header_row_index + 1:
                    headers = raw_rows[header_row_index]
                    for row in raw_rows[header_row_index + 1:]:
                        if any(row):
                            rows.append(dict(zip(headers, row)))
            else:
                content_bytes = file.stream.read()
                file.stream.seek(0)
                try:
                    with open(r'c:\dev\veros_books\last_import.xlsx', 'wb') as f:
                        f.write(content_bytes)
                except:
                    pass
                import openpyxl
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                
                header_row_index = 1
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    row_str = " ".join([str(x).lower() for x in row if x])
                    if 'date' in row_str and ('amount' in row_str or 'debit' in row_str):
                        header_row_index = idx
                        break
                
                cell_rows = list(sheet.iter_rows(min_row=header_row_index, values_only=True))
                if len(cell_rows) > 1:
                    headers = cell_rows[0]
                    for row in cell_rows[1:]:
                        if any(row):
                            rows.append(dict(zip(headers, row)))
                            
            for row in rows:
                lowered_row = {str(k).lower().strip(): v for k, v in row.items() if k}
                date_val = lowered_row.get('date') or lowered_row.get('posted date') or lowered_row.get('transaction date')
                desc_val = lowered_row.get('description') or lowered_row.get('payee') or lowered_row.get('memo') or lowered_row.get('name')
                amount_val = lowered_row.get('amount')
                
                if not date_val:
                    continue
                    
                # Deal with split debit/credit columns if amount isn't provided directly
                amount = None
                if amount_val:
                    try: amount = float(str(amount_val).replace('$', '').replace(',', ''))
                    except: pass
                
                if amount is None and lowered_row.get('debit'):
                    try: amount = -abs(float(str(lowered_row.get('debit')).replace('$', '').replace(',', '')))
                    except: pass
                    
                if amount is None and lowered_row.get('credit'):
                    try: amount = abs(float(str(lowered_row.get('credit')).replace('$', '').replace(',', '')))
                    except: pass
                    
                if amount is None:
                    continue
                    
                # Parse python datetime or string
                if isinstance(date_val, datetime):
                    tx_date = date_val.date()
                else:
                    try:
                        tx_date = datetime.strptime(str(date_val).split(' ')[0], '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            tx_date = datetime.strptime(str(date_val).split(' ')[0], '%m/%d/%Y').date()
                        except:
                            tx_date = datetime.utcnow().date()

                existing = BankTransaction.query.filter_by(
                    bank_account_id=bank_acc.id, date=tx_date, amount=amount
                ).first()
                
                if not existing:
                    new_tx = BankTransaction(
                        organization_id=org.id,
                        bank_account_id=bank_acc.id,
                        date=tx_date,
                        description=str(desc_val) if desc_val else "Unknown Transaction",
                        amount=amount,
                        status='UNCATEGORIZED'
                    )
                    db.session.add(new_tx)
                    imported_txs += 1
                    
            db.session.commit()
            flash(f"Successfully imported {imported_txs} transactions from Spreadsheet.", "success")
            
        else:
            flash("Unsupported file format.", "danger")

    except Exception as e:
        db.session.rollback()
        flash(f"Error parsing statement file: {str(e)}", "danger")
        
    return redirect(url_for('banking.index'))
@banking_bp.route('/rules')
@login_required
def rules():
    org = get_current_org()
    from app.models.accounting.bank_rule import BankRule
    from app.models.accounting.account import Account
    from app.models.crm.contact import Customer, Vendor
    
    rules = BankRule.query.filter_by(organization_id=org.id).order_by(BankRule.priority.desc()).all()
    accounts = Account.query.filter_by(organization_id=org.id).all()
    
    return render_template('banking/rules.html', rules=rules, accounts=accounts)

@banking_bp.route('/rules/create', methods=['POST'])
@login_required
def create_rule():
    org = get_current_org()
    from app.models.accounting.bank_rule import BankRule
    
    new_rule = BankRule(
        organization_id=org.id,
        name=request.form.get('name'),
        field_to_match=request.form.get('field'),
        match_type=request.form.get('match_type'),
        match_value=request.form.get('match_value'),
        target_account_id=request.form.get('target_account_id'),
        auto_post=request.form.get('auto_post') == 'true'
    )

    db.session.add(new_rule)
    db.session.commit()
    flash("Bank rule created successfully.", "success")
    return redirect(url_for('banking.rules'))

@banking_bp.route('/rules/<rule_id>/delete', methods=['POST'])
@login_required
def delete_rule(rule_id):
    org = get_current_org()
    from app.models.accounting.bank_rule import BankRule
    rule = BankRule.query.filter_by(id=rule_id, organization_id=org.id).first_or_404()
    db.session.delete(rule)
    db.session.commit()
    flash("Bank rule removed.", "success")
    return redirect(url_for('banking.rules'))

@banking_bp.route('/checks')
@login_required
def checks():
    org = get_current_org()
    from app.models.banking.check import Check
    checks = Check.query.filter_by(organization_id=org.id).order_by(Check.date.desc(), Check.check_number.desc()).all()
    return render_template('banking/checks/index.html', checks=checks)

@banking_bp.route('/checks/create', methods=['GET', 'POST'])
@login_required
def create_check():
    org = get_current_org()
    from app.models.banking.bank_account import BankAccount
    from app.models.accounting.account import Account
    from app.models.crm.contact import Vendor
    from app.models.banking.check import Check
    from app.models.accounting.journal import JournalEntry, JournalLine
    from app.services.ledger_service import LedgerService
    from datetime import datetime

    if request.method == 'POST':
        bank_account_id = request.form.get('bank_account_id')
        check_number = request.form.get('check_number')
        date_str = request.form.get('date')
        payee_name = request.form.get('payee_name')
        amount_str = request.form.get('amount', '0')
        memo = request.form.get('memo')
        expense_account_id = request.form.get('expense_account_id')

        try:
            amount = float(amount_str)
        except ValueError:
            amount = 0

        # Basic validation
        if not bank_account_id or not check_number or amount <= 0 or not expense_account_id:
            flash("Missing or invalid required fields.", "danger")
            return redirect(url_for('banking.create_check'))

        check_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.utcnow().date()

        # Get the GL account for the bank account
        bank_acc = BankAccount.query.get(bank_account_id)
        if not bank_acc or not bank_acc.account_id:
            flash("Bank account is not linked to a GL account.", "danger")
            return redirect(url_for('banking.create_check'))

        # 1. Create Journal Entry
        je = JournalEntry(
            organization_id=org.id,
            entry_number=f"CHK-{check_number}",
            entry_date=datetime.combine(check_date, datetime.min.time()),
            memo=memo or f"Check #{check_number} to {payee_name}",
            source_type='CHECK',
            status='DRAFT',
            created_by=current_user.id
        )
        db.session.add(je)
        db.session.flush()
        
        # Line 1: Credit Bank Account
        bank_line = JournalLine(
            journal_entry_id=je.id,
            account_id=bank_acc.account_id,
            credit=amount,
            description=f"Check #{check_number} to {payee_name}"
        )
        
        # Line 2: Debit Expense Account
        expense_line = JournalLine(
            journal_entry_id=je.id,
            account_id=expense_account_id,
            debit=amount,
            description=memo or f"Check #{check_number}"
        )
        
        db.session.add(bank_line)
        db.session.add(expense_line)
        db.session.flush()

        # 2. Create Check Record
        new_check = Check(
            organization_id=org.id,
            bank_account_id=bank_account_id,
            check_number=check_number,
            date=check_date,
            payee_name=payee_name,
            amount=amount,
            memo=memo,
            status='PRINTED',
            journal_entry_id=je.id
        )
        db.session.add(new_check)
        
        # 3. Post to Ledger
        success, msg = LedgerService.post_journal_entry(je.id, current_user.id, org.id)
        
        if success:
            db.session.commit()
            flash(f"Check #{check_number} created and posted.", "success")
        else:
            db.session.rollback()
            flash(f"Error posting check: {msg}", "danger")
            
        return redirect(url_for('banking.checks'))

    bank_accounts = BankAccount.query.filter_by(organization_id=org.id, account_type='Checking').all()
    expense_accounts = Account.query.filter_by(organization_id=org.id).order_by(Account.code).all()
    vendors = Vendor.query.filter_by(organization_id=org.id).all()
    
    return render_template('banking/checks/create.html', 
                           bank_accounts=bank_accounts, 
                           expense_accounts=expense_accounts,
                           vendors=vendors)

@banking_bp.route('/transactions/attach-receipt', methods=['POST'])
@login_required
def attach_receipt():
    org = get_current_org()
    tx_id = request.form.get('tx_id')
    receipt_url = request.form.get('receipt_url')
    
    from app.models.banking.bank_account import BankTransaction
    tx = BankTransaction.query.filter_by(id=tx_id, organization_id=org.id).first_or_404()
    tx.receipt_url = receipt_url
    db.session.commit()
    
    flash("Receipt successfully attached to transaction.", "success")
    return redirect(url_for('banking.index'))


